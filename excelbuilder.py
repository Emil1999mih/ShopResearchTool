import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
def save_to_excel(data, filepath): #salvarea datelor in  format excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Extracted Data'

    headers = [key.capitalize() for key in data[0].keys()]
    sheet.append(headers)

    for row in data:
        sheet.append(list(row.values()))

    last_row = len(data) + 1
    last_column = len(headers)

    table_ref = f"A1:{sheet.cell(row=last_row, column=last_column).coordinate}"

    table = Table(displayName="Products", ref=table_ref)

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style
    for column in sheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    sheet.add_table(table)
    workbook.save(filepath)


def add_price_analysis(data, filepath):
    workbook = openpyxl.load_workbook(filepath)

    data_sheet = workbook['Extracted Data']

    analysis_sheet = workbook.create_sheet(title='Price Analysis')

    headers = [cell.value for cell in data_sheet[1]]
    try:
        price_col_idx = headers.index('Price') + 1
        title_col_idx = headers.index('Title') + 1
    except ValueError:
        print("No 'Price' or 'Title' column found in the data")
        return

    prices = []
    price_details = []
    for row in range(2, data_sheet.max_row + 1):
        price_cell = data_sheet.cell(row=row, column=price_col_idx).value
        title_cell = data_sheet.cell(row=row, column=title_col_idx).value
        try:
            if isinstance(price_cell, str):
                price_clean = price_cell.replace('$', '').replace('€', '').replace(',', '').replace(' ', '')
                price = float(price_clean)
            else:
                price = float(price_cell)
            prices.append(price)
            price_details.append({'price': price, 'title': title_cell})
        except (ValueError, TypeError):
            continue

    if prices:
        average_price = sum(prices) / len(prices)
        highest_price = max(prices)
        lowest_price = min(prices)
        price_count = len(prices)

        # Find products with highest and lowest prices
        highest_price_products = [item['title'] for item in price_details if item['price'] == highest_price]
        lowest_price_products = [item['title'] for item in price_details if item['price'] == lowest_price]
    else:
        average_price = highest_price = lowest_price = price_count = 0
        highest_price_products = lowest_price_products = []

    analysis_sheet.append(["Price Analysis"])
    analysis_sheet.append([])
    analysis_sheet.append(["Metric", "Value"])
    analysis_sheet.append(["Number of Products with Price", price_count])
    analysis_sheet.append(["Average Price", f"{average_price:.2f}"])
    analysis_sheet.append(["Highest Price", f"{highest_price:.2f}"])
    analysis_sheet.append(["Highest Price Products", ", ".join(highest_price_products)])
    analysis_sheet.append(["Lowest Price", f"{lowest_price:.2f}"])
    analysis_sheet.append(["Lowest Price Products", ", ".join(lowest_price_products)])

    for row in analysis_sheet.iter_rows(min_row=1, max_row=7, min_col=1, max_col=2):
        for cell in row:
            cell.style = 'Normal'
            if row[0].row == 1:
                cell.font = openpyxl.styles.Font(bold=True, size=14)
            elif row[0].row == 3:
                cell.font = openpyxl.styles.Font(bold=True)

    for column in analysis_sheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        analysis_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    workbook.save(filepath)

def add_price_ranges(data, filepath):

    workbook = openpyxl.load_workbook(filepath)

    data_sheet = workbook['Extracted Data']

    analysis_sheet = workbook.create_sheet(title='Price Ranges')
    ranges = [
        (0, 99.99, 'Under 100'),
        (100, 200, '100-200'),
        (200, float('inf'), 'Over 200')
    ]

    range_data = {range_name: [] for _, _, range_name in ranges}

    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        title = row[0]
        price = float(row[2])

        for min_price, max_price, range_name in ranges:
            if min_price <= price <= max_price:
                range_data[range_name].append((title, price))
                break

    analysis_sheet['A1'] = 'Price Range Analysis'
    analysis_sheet['A1'].font = Font(bold=True, size=14)

    current_row = 3

    for _, _, range_name in ranges:

        analysis_sheet[f'A{current_row}'] = range_name
        analysis_sheet[f'A{current_row}'].font = Font(bold=True)
        analysis_sheet[f'B{current_row}'] = f'Total Items: {len(range_data[range_name])}'
        current_row += 1

        analysis_sheet[f'A{current_row}'] = 'Product Name'
        analysis_sheet[f'B{current_row}'] = 'Price'
        analysis_sheet[f'A{current_row}'].font = Font(bold=True)
        analysis_sheet[f'B{current_row}'].font = Font(bold=True)
        current_row += 1

        for title, price in range_data[range_name]:
            analysis_sheet[f'A{current_row}'] = title
            analysis_sheet[f'B{current_row}'] = price
            current_row += 1

        current_row += 1

    analysis_sheet.column_dimensions['A'].width = 60
    analysis_sheet.column_dimensions['B'].width = 15

    workbook.save(filepath)



def add_graphs(data, filepath):
    workbook = openpyxl.load_workbook(filepath)

    chart_sheet = workbook.create_sheet(title='Graphs Analysis')

    range_sheet = workbook['Price Ranges']
    ranges = [
        (0, 99.99, 'Under 100'),
        (100, 200, '100-200'),
        (200, float('inf'), 'Over 200')
    ]

    range_counts = {range_name: 0 for _, _, range_name in ranges}

    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        price = float(row[2])
        for min_price, max_price, range_name in ranges:
            if min_price <= price <= max_price:
                range_counts[range_name] += 1
                break

    chart_sheet['A1'] = 'Price Range Distribution'
    chart_sheet['A2'] = 'Price Range'
    chart_sheet['B2'] = 'Number of Products'

    for i, (range_name, count) in enumerate(range_counts.items(), start=3):
        chart_sheet[f'A{i}'] = range_name
        chart_sheet[f'B{i}'] = count

    pie = PieChart()
    pie.title = "Product Distribution by Price Range"

    data = Reference(chart_sheet, min_col=2, min_row=2, max_row=5)
    labels = Reference(chart_sheet, min_col=1, min_row=3, max_row=5)

    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)

    pie.height = 20
    pie.width = 20

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True

    slice_colors = ['87CEEB', '4682B4', '000080']
    for i in range(len(range_counts)):
        slice = DataPoint(idx=i)
        slice.graphicalProperties.solidFill = slice_colors[i]
        pie.series[0].data_points.append(slice)

    chart_sheet.add_chart(pie, "D2")

    chart_sheet.column_dimensions['A'].width = 15
    chart_sheet.column_dimensions['B'].width = 20

    workbook.save(filepath)

